# ============================================================
# services/product_scoring.py
# ============================================================
# RESPONSABILIDAD:
# Servicio para aplicar el score comercial de NIA a productos
# reales del catálogo MongoDB.
#
# Este módulo:
# - Lee registros normalizados de scoring.
# - Valida CODIGO y score.
# - Busca productos reales en products_catalog.
# - Actualiza campos de scoring en MongoDB.
#
# IMPORTANTE:
# - NO recomienda productos.
# - NO modifica el orquestador.
# - NO borra productos.
# - NO inventa score.
# - SOLO actualiza productos existentes por CODIGO.
#
# Regla de Don Andrés:
# - products_catalog = fuente de verdad para productos reales.
# - score_nia = prioridad comercial para ordenar recomendaciones.
# ============================================================

from __future__ import annotations

import logging
import math
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from pymongo import UpdateOne

from .mongo import get_collection

logger = logging.getLogger(__name__)


# ============================================================
# CONFIGURACIÓN
# ============================================================

DEFAULT_CODE_COLUMN = "CODIGO"
DEFAULT_SCORE_COLUMN = "score"

SCORE_FIELD = "score_nia"
SCORE_SOURCE_FIELD = "score_source"
SCORE_VERSION_FIELD = "score_version"
SCORE_UPDATED_AT_FIELD = "score_updated_at"

DEFAULT_SCORE_SOURCE = "excel_don_andres"
DEFAULT_BATCH_SIZE = 1000


# ============================================================
# UTILIDADES
# ============================================================

def _now_utc() -> datetime:
    """
    Retorna fecha/hora actual en UTC.
    """
    return datetime.now(timezone.utc)


def _safe_str(value: Any) -> str:
    """
    Convierte cualquier valor a string limpio.
    """
    if value is None:
        return ""

    try:
        return str(value).strip()
    except Exception:
        return ""


def normalize_code(value: Any) -> str:
    """
    Normaliza el código del producto.

    Casos:
    - " P382280 " -> "P382280"
    - 123456.0 -> "123456"
    - "123456.0" -> "123456"

    No cambia mayúsculas/minúsculas de manera destructiva,
    pero para códigos VIA se usa uppercase por consistencia.
    """
    text = _safe_str(value)

    if not text:
        return ""

    # Si Excel leyó un código numérico como float, puede quedar "123456.0".
    if text.endswith(".0"):
        possible_int = text[:-2]
        if possible_int.isdigit():
            text = possible_int

    return text.strip().upper()


def parse_score(value: Any) -> Optional[float]:
    """
    Convierte el score del Excel a float válido.

    Retorna None si:
    - está vacío
    - no es numérico
    - es NaN o infinito
    """
    if value is None:
        return None

    try:
        if isinstance(value, str):
            value = value.strip().replace(",", ".")

        score = float(value)

        if math.isnan(score) or math.isinf(score):
            return None

        return score

    except Exception:
        return None


def _chunk_list(items: List[Any], size: int) -> List[List[Any]]:
    """
    Divide una lista en bloques.
    """
    if size <= 0:
        size = DEFAULT_BATCH_SIZE

    return [
        items[index:index + size]
        for index in range(0, len(items), size)
    ]


def _find_column_index(
    headers: List[Any],
    expected_name: str,
) -> Optional[int]:
    """
    Busca una columna por nombre sin depender de mayúsculas/minúsculas.
    """
    expected = _safe_str(expected_name).lower()

    for index, header in enumerate(headers):
        current = _safe_str(header).lower()

        if current == expected:
            return index

    return None


# ============================================================
# LECTURA DEL EXCEL
# ============================================================

def read_scoring_excel(
    excel_path: str | Path,
    code_column: str = DEFAULT_CODE_COLUMN,
    score_column: str = DEFAULT_SCORE_COLUMN,
    max_rows: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Lee el Excel de scoring compartido por Don Andrés.

    Retorna:
    {
        "ok": True,
        "records": [
            {"codigo": "P382280", "score": 88.4, "row_number": 2}
        ],
        "summary": {...},
        "errors": []
    }

    Nota:
    Usa openpyxl en modo read_only para soportar archivos grandes.
    """
    path = Path(excel_path)

    if not path.exists():
        return {
            "ok": False,
            "records": [],
            "summary": {
                "total_rows_read": 0,
                "valid_records": 0,
                "invalid_rows": 0,
                "duplicate_codes": 0,
            },
            "errors": [f"No existe el archivo Excel: {path}"],
        }

    if not path.is_file():
        return {
            "ok": False,
            "records": [],
            "summary": {
                "total_rows_read": 0,
                "valid_records": 0,
                "invalid_rows": 0,
                "duplicate_codes": 0,
            },
            "errors": [f"La ruta no corresponde a un archivo: {path}"],
        }

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return {
            "ok": False,
            "records": [],
            "summary": {
                "total_rows_read": 0,
                "valid_records": 0,
                "invalid_rows": 0,
                "duplicate_codes": 0,
            },
            "errors": [
                "No está instalada la librería openpyxl. Instala con: pip install openpyxl",
                str(exc),
            ],
        }

    records: List[Dict[str, Any]] = []
    errors: List[str] = []
    invalid_rows = 0
    total_rows_read = 0

    try:
        workbook = load_workbook(
            filename=str(path),
            read_only=True,
            data_only=True,
        )

        sheet = workbook.active

        rows_iterator = sheet.iter_rows(values_only=True)

        try:
            headers = list(next(rows_iterator))
        except StopIteration:
            return {
                "ok": False,
                "records": [],
                "summary": {
                    "total_rows_read": 0,
                    "valid_records": 0,
                    "invalid_rows": 0,
                    "duplicate_codes": 0,
                },
                "errors": ["El Excel está vacío."],
            }

        code_index = _find_column_index(headers, code_column)
        score_index = _find_column_index(headers, score_column)

        if code_index is None:
            return {
                "ok": False,
                "records": [],
                "summary": {
                    "total_rows_read": 0,
                    "valid_records": 0,
                    "invalid_rows": 0,
                    "duplicate_codes": 0,
                },
                "errors": [f"No se encontró la columna requerida: {code_column}"],
            }

        if score_index is None:
            return {
                "ok": False,
                "records": [],
                "summary": {
                    "total_rows_read": 0,
                    "valid_records": 0,
                    "invalid_rows": 0,
                    "duplicate_codes": 0,
                },
                "errors": [f"No se encontró la columna requerida: {score_column}"],
            }

        seen_codes = set()
        duplicate_codes = 0

        for excel_row_number, row in enumerate(rows_iterator, start=2):
            if max_rows is not None and total_rows_read >= max_rows:
                break

            total_rows_read += 1

            row_values = list(row)

            raw_code = row_values[code_index] if code_index < len(row_values) else None
            raw_score = row_values[score_index] if score_index < len(row_values) else None

            codigo = normalize_code(raw_code)
            score = parse_score(raw_score)

            if not codigo or score is None:
                invalid_rows += 1
                continue

            if codigo in seen_codes:
                duplicate_codes += 1

            seen_codes.add(codigo)

            records.append({
                "codigo": codigo,
                "score": score,
                "row_number": excel_row_number,
            })

        workbook.close()

        return {
            "ok": True,
            "records": records,
            "summary": {
                "excel_path": str(path),
                "sheet_name": sheet.title,
                "total_rows_read": total_rows_read,
                "valid_records": len(records),
                "invalid_rows": invalid_rows,
                "duplicate_codes": duplicate_codes,
                "code_column": code_column,
                "score_column": score_column,
            },
            "errors": errors,
        }

    except Exception as error:
        logger.exception("Error leyendo Excel de scoring")

        return {
            "ok": False,
            "records": records,
            "summary": {
                "excel_path": str(path),
                "total_rows_read": total_rows_read,
                "valid_records": len(records),
                "invalid_rows": invalid_rows,
                "duplicate_codes": 0,
            },
            "errors": [str(error)],
        }


# ============================================================
# VALIDACIÓN CONTRA MONGODB
# ============================================================

def find_existing_product_codes(
    codes: List[str],
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> set[str]:
    """
    Busca qué códigos del Excel existen realmente en products_catalog.

    Esto NO modifica MongoDB.
    """
    clean_codes = [
        normalize_code(code)
        for code in codes
        if normalize_code(code)
    ]

    existing_codes: set[str] = set()

    if not clean_codes:
        return existing_codes

    col = get_collection()

    for code_batch in _chunk_list(clean_codes, batch_size):
        cursor = col.find(
            {"CODIGO": {"$in": code_batch}},
            {"_id": 0, "CODIGO": 1},
        )

        for doc in cursor:
            codigo = normalize_code(doc.get("CODIGO"))
            if codigo:
                existing_codes.add(codigo)

    return existing_codes


def preview_scoring_records(
    records: List[Dict[str, Any]],
    limit: int = 10,
) -> List[Dict[str, Any]]:
    """
    Devuelve una muestra corta de registros de scoring.
    """
    if limit <= 0:
        limit = 10

    return records[:limit]


# ============================================================
# ACTUALIZACIÓN DE SCORE EN MONGODB
# ============================================================

def apply_product_scores(
    records: List[Dict[str, Any]],
    dry_run: bool = True,
    score_version: Optional[str] = None,
    score_source: str = DEFAULT_SCORE_SOURCE,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> Dict[str, Any]:
    """
    Aplica score_nia a products_catalog.

    Si dry_run=True:
    - NO actualiza MongoDB.
    - Solo valida coincidencias por CODIGO.

    Si dry_run=False:
    - Actualiza productos existentes por CODIGO.
    - No inserta productos nuevos.
    - No borra nada.

    Regla para duplicados:
    - Si un mismo CODIGO aparece varias veces en el Excel,
      se conserva el score más alto.
    - Esto evita múltiples updates al mismo producto.
    - No inventa score: siempre toma un valor existente del Excel.
    """
    if not isinstance(records, list):
        return {
            "ok": False,
            "dry_run": dry_run,
            "summary": {},
            "errors": ["records debe ser una lista."],
        }

    # --------------------------------------------------------
    # 1. Validar registros base
    # --------------------------------------------------------
    valid_records = []

    for item in records:
        if not isinstance(item, dict):
            continue

        codigo = normalize_code(item.get("codigo"))
        score = parse_score(item.get("score"))

        if not codigo or score is None:
            continue

        valid_records.append({
            "codigo": codigo,
            "score": score,
            "row_number": item.get("row_number"),
        })

    # --------------------------------------------------------
    # 2. Deduplicar por CODIGO
    # --------------------------------------------------------
    # Si hay códigos repetidos, dejamos el score más alto.
    # Esto es importante antes del bulk_write para no actualizar
    # varias veces el mismo documento en una sola ejecución.
    deduped_by_code: Dict[str, Dict[str, Any]] = {}
    duplicate_codes_seen = set()

    for item in valid_records:
        codigo = item["codigo"]

        if codigo in deduped_by_code:
            duplicate_codes_seen.add(codigo)

            current_score = deduped_by_code[codigo]["score"]
            new_score = item["score"]

            if new_score > current_score:
                deduped_by_code[codigo] = item
        else:
            deduped_by_code[codigo] = item

    deduped_records = list(deduped_by_code.values())

    # --------------------------------------------------------
    # 3. Validar cuáles existen realmente en MongoDB
    # --------------------------------------------------------
    codes = [item["codigo"] for item in deduped_records]
    existing_codes = find_existing_product_codes(codes, batch_size=batch_size)

    matched_records = [
        item for item in deduped_records
        if item["codigo"] in existing_codes
    ]

    missing_records = [
        item for item in deduped_records
        if item["codigo"] not in existing_codes
    ]

    # --------------------------------------------------------
    # 4. Dry-run: no modifica MongoDB
    # --------------------------------------------------------
    if dry_run:
        return {
            "ok": True,
            "dry_run": True,
            "summary": {
                "input_records": len(records),
                "valid_records": len(valid_records),
                "unique_codes_after_dedup": len(deduped_records),
                "duplicate_codes_collapsed": len(duplicate_codes_seen),
                "matched_products": len(matched_records),
                "missing_products": len(missing_records),
                "would_update": len(matched_records),
                "modified_count": 0,
                "score_field": SCORE_FIELD,
                "score_source": score_source,
                "score_version": score_version,
                "dedup_strategy": "max_score_per_codigo",
            },
            "samples": {
                "matched": matched_records[:10],
                "missing": missing_records[:10],
            },
            "errors": [],
        }

    # --------------------------------------------------------
    # 5. Aplicación real en MongoDB
    # --------------------------------------------------------
    col = get_collection()

    if not score_version:
        score_version = _now_utc().strftime("%Y-%m-%d")

    updated_at = _now_utc()
    total_modified = 0
    total_matched_bulk = 0
    batch_errors: List[str] = []

    for record_batch in _chunk_list(matched_records, batch_size):
        operations = []

        for item in record_batch:
            operations.append(
                UpdateOne(
                    {"CODIGO": item["codigo"]},
                    {
                        "$set": {
                            SCORE_FIELD: item["score"],
                            SCORE_SOURCE_FIELD: score_source,
                            SCORE_VERSION_FIELD: score_version,
                            SCORE_UPDATED_AT_FIELD: updated_at,
                        }
                    },
                    upsert=False,
                )
            )

        if not operations:
            continue

        try:
            result = col.bulk_write(
                operations,
                ordered=False,
            )

            total_modified += result.modified_count
            total_matched_bulk += result.matched_count

        except Exception as error:
            logger.exception("Error aplicando batch de scoring")
            batch_errors.append(str(error))

    return {
        "ok": len(batch_errors) == 0,
        "dry_run": False,
        "summary": {
            "input_records": len(records),
            "valid_records": len(valid_records),
            "unique_codes_after_dedup": len(deduped_records),
            "duplicate_codes_collapsed": len(duplicate_codes_seen),
            "matched_products": len(matched_records),
            "missing_products": len(missing_records),
            "bulk_matched_count": total_matched_bulk,
            "modified_count": total_modified,
            "score_field": SCORE_FIELD,
            "score_source": score_source,
            "score_version": score_version,
            "score_updated_at": updated_at.isoformat(),
            "dedup_strategy": "max_score_per_codigo",
        },
        "samples": {
            "matched": matched_records[:10],
            "missing": missing_records[:10],
        },
        "errors": batch_errors,
    }

# ============================================================
# FUNCIÓN DE ALTO NIVEL
# ============================================================

def apply_scoring_from_excel(
    excel_path: str | Path,
    dry_run: bool = True,
    max_rows: Optional[int] = None,
    score_version: Optional[str] = None,
    score_source: str = DEFAULT_SCORE_SOURCE,
    code_column: str = DEFAULT_CODE_COLUMN,
    score_column: str = DEFAULT_SCORE_COLUMN,
    batch_size: int = DEFAULT_BATCH_SIZE,
) -> Dict[str, Any]:
    """
    Flujo completo:
    1. Lee Excel.
    2. Valida CODIGO + score.
    3. Hace dry-run o actualización real en MongoDB.
    """
    read_result = read_scoring_excel(
        excel_path=excel_path,
        code_column=code_column,
        score_column=score_column,
        max_rows=max_rows,
    )

    if not read_result.get("ok"):
        return {
            "ok": False,
            "stage": "read_excel",
            "read_result": read_result,
            "apply_result": None,
            "errors": read_result.get("errors", []),
        }

    apply_result = apply_product_scores(
        records=read_result.get("records", []),
        dry_run=dry_run,
        score_version=score_version,
        score_source=score_source,
        batch_size=batch_size,
    )

    return {
        "ok": apply_result.get("ok", False),
        "stage": "dry_run" if dry_run else "apply",
        "read_result": {
            "summary": read_result.get("summary", {}),
            "errors": read_result.get("errors", []),
            "sample": preview_scoring_records(read_result.get("records", []), limit=5),
        },
        "apply_result": apply_result,
        "errors": apply_result.get("errors", []),
    }